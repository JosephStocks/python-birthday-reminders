from datetime import datetime
from typing import Optional

from openpyxl import load_workbook
from pydantic import BaseModel


class RawDataEntry(BaseModel):
    event_type: str
    first_name: str
    last_name: Optional[str]
    date: Optional[datetime]
    sure_about_year: str
    sex: str
    family_or_friend: str
    side_of_family: str
    relation: str


def load_excel_data(filename: str | None) -> list[RawDataEntry]:
    if not filename:
        raise ValueError("A filename must be provided.")

    workbook = load_workbook(filename)
    sheet = workbook.worksheets[0]

    header = [str(cell.value) for cell in sheet[1]]

    data: list[RawDataEntry] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        zipped_dict = dict(zip(header, row))
        entry = RawDataEntry.model_validate(zipped_dict)
        data.append(entry)

    return data
