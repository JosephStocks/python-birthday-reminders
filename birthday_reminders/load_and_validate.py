import json
from dataclasses import dataclass
from datetime import date, datetime
from enum import StrEnum, auto
from typing import Optional

from openpyxl import load_workbook
from pydantic import BaseModel


class RawDataEntry(BaseModel):
    event_type: str
    first_name: str
    last_name: Optional[str]
    date: Optional[datetime]
    sure_about_year: str
    sex: str
    family_or_friend: str
    side_of_family: str
    relation: str
    days_ahead_alert: Optional[str]


class FamilyOrFriend(StrEnum):
    family = auto()
    friend = auto()


class SpecificRelation(StrEnum):
    spouse = auto()
    child = auto()
    sibling = auto()
    parent = auto()
    grandparent = auto()
    sibling_in_law = auto()
    niece_or_nephew = auto()
    friend = auto()


class Sex(StrEnum):
    m = auto()
    f = auto()


class SideOfFamily(StrEnum):
    joe = auto()
    katie = auto()
    core = auto()


@dataclass
class Birthday:
    name: str
    date: date
    sure_about_year: bool
    sex: Sex
    family_or_friend: FamilyOrFriend
    side_of_family: SideOfFamily
    relation: SpecificRelation
    days_ahead_alert: Optional[list[int]]


def load_excel_data(filename: str) -> list[RawDataEntry]:
    if not filename:
        raise ValueError("A filename must be provided.")

    workbook = load_workbook(filename)
    sheet = workbook.worksheets[0]

    header = [str(cell.value) for cell in sheet[1]]

    data: list[RawDataEntry] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        zipped_dict = dict(zip(header, row))
        entry = RawDataEntry.model_validate(zipped_dict)
        data.append(entry)

    return data


def process_birthdays(raw_data: list[RawDataEntry]) -> list[Birthday]:
    birthdays: list[Birthday] = []
    for entry in raw_data:
        if entry.event_type == "birthday" and entry.date is not None:
            birthday = Birthday(
                name=f"{entry.first_name} {entry.last_name}".strip(),
                date=entry.date.date(),
                sure_about_year=entry.sure_about_year == "yes",
                sex=Sex[entry.sex],
                family_or_friend=FamilyOrFriend[entry.family_or_friend],
                side_of_family=SideOfFamily[entry.side_of_family],
                relation=SpecificRelation[entry.relation],
                days_ahead_alert=(
                    json.loads(entry.days_ahead_alert)
                    if entry.days_ahead_alert
                    else None
                ),
            )
            birthdays.append(birthday)
    return birthdays
