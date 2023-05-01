from openpyxl import load_workbook


def load_excel_data(filename):
    # Open up the Excel file.
    workbook = load_workbook(filename)

    # Get the first sheet.
    sheet = workbook.worksheets[0]

    # Get the header row (column names)
    header = [cell.value for cell in sheet[1]]

    # Convert the sheet to a list of dictionaries
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(header, row)))

    return data
